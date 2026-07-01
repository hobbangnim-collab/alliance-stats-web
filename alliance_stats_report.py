import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


DEFAULT_INPUT_DIR = "."
DEFAULT_OUTPUT = "alliance_stats_report.xlsx"
DEFAULT_MEMBERS_FILE = "members.txt"
DEFAULT_OUTPUT_DIR = "결과물"

KEY_COLUMN = "회원"
GROUP_COLUMN = "그룹"
STATE_COLUMN = "소속 주"
RANK_COLUMN = "공헌 랭킹"

COLUMN_ALIASES = {
    "동맹 공헌 랭킹": RANK_COLUMN,
    "소속 진영": STATE_COLUMN,
}

NUMERIC_COLUMNS = [
    "공헌 랭킹",
    "공헌이번 주",
    "전공이번 주",
    "협공이번 주",
    "기부이번 주",
    "공헌총량",
    "전공총량",
    "협공총량",
    "기부총량",
    "세력치",
]

DELTA_COLUMNS = [
    "세력치",
    "공헌총량",
    "전공총량",
    "협공총량",
    "기부총량",
    "공헌 랭킹",
]

TOP_COLUMNS = [
    "세력치",
    "공헌총량",
    "전공총량",
    "공헌 랭킹",
]

DATE_RE = re.compile(r"(\d{4})-(\d{2})-(\d{2})\s+(\d{4})")
CHINESE_DATE_RE = re.compile(r"(\d{4})年(\d{2})月(\d{2})日(\d{2})时(\d{2})分(?:(\d{2})秒)?")


def parse_snapshot_time(path):
    match = DATE_RE.search(path.stem)
    if match:
        year, month, day, hm = match.groups()
        hour = int(hm[:2])
        minute = int(hm[2:])
        return datetime(int(year), int(month), int(day), hour, minute)

    match = CHINESE_DATE_RE.search(path.stem)
    if match:
        year, month, day, hour, minute, second = match.groups()
        return datetime(int(year), int(month), int(day), int(hour), int(minute), int(second or 0))

    return datetime.fromtimestamp(path.stat().st_mtime)


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, skipinitialspace=True)
        rows = {}
        for raw in reader:
            row = {key.strip(): value.strip() for key, value in raw.items() if key}
            for alias, canonical in COLUMN_ALIASES.items():
                if alias in row and (canonical not in row or row.get(canonical, "") == ""):
                    row[canonical] = row[alias]
            member = row.get(KEY_COLUMN)
            if not member:
                continue
            for column in NUMERIC_COLUMNS:
                row[column] = parse_number(row.get(column))
            rows[member] = row
        return rows


def parse_number(value):
    if value is None or value == "":
        return 0
    return int(str(value).replace(",", "").strip())


def load_snapshots(input_dir):
    files = sorted(Path(input_dir).glob("*.csv"), key=parse_snapshot_time)
    if len(files) < 2:
        raise RuntimeError("비교하려면 CSV 파일이 최소 2개 필요합니다.")

    snapshots = []
    for path in files:
        snapshots.append(
            {
                "path": path,
                "time": parse_snapshot_time(path),
                "label": parse_snapshot_time(path).strftime("%Y-%m-%d %H:%M"),
                "rows": read_csv(path),
            }
        )
    return snapshots


def build_comparison(previous, current, member_filter=None):
    if member_filter:
        members = sorted(member_filter)
    else:
        members = sorted(set(previous["rows"]) | set(current["rows"]))
    result = []

    for member in members:
        old = previous["rows"].get(member)
        new = current["rows"].get(member)

        if old and new:
            status = "유지"
            display = new
        elif new:
            status = "신규"
            display = new
        else:
            status = "이탈"
            display = old or {}

        row = {
            "상태": status,
            KEY_COLUMN: member,
            GROUP_COLUMN: display.get(GROUP_COLUMN, ""),
            STATE_COLUMN: display.get(STATE_COLUMN, ""),
        }

        for column in DELTA_COLUMNS:
            old_value = old.get(column, 0) if old else 0
            new_value = new.get(column, 0) if new else 0
            delta = new_value - old_value
            if column == RANK_COLUMN and old and new:
                delta = old_value - new_value

            row[f"이전 {column}"] = old_value
            row[f"현재 {column}"] = new_value
            row[f"{column} 변화"] = delta

        result.append(row)

    return result


def build_group_summary(comparison):
    groups = defaultdict(lambda: defaultdict(int))

    for row in comparison:
        group = row.get(GROUP_COLUMN) or "그룹 없음"
        groups[group]["인원"] += 1
        if row["상태"] == "신규":
            groups[group]["신규"] += 1
        elif row["상태"] == "이탈":
            groups[group]["이탈"] += 1

        for column in DELTA_COLUMNS:
            groups[group][f"{column} 변화"] += row[f"{column} 변화"]
            groups[group][f"현재 {column}"] += row[f"현재 {column}"]

    rows = []
    for group, values in groups.items():
        item = {"그룹": group}
        item.update(values)
        rows.append(item)

    rows.sort(key=lambda value: value.get("세력치 변화", 0), reverse=True)
    return rows


def build_trend(snapshots):
    members = sorted({member for snapshot in snapshots for member in snapshot["rows"]})
    rows = []
    for member in members:
        row = {KEY_COLUMN: member}
        latest_group = ""
        for snapshot in snapshots:
            source = snapshot["rows"].get(member)
            row[snapshot["label"]] = source.get("세력치", 0) if source else 0
            if source:
                latest_group = source.get(GROUP_COLUMN, latest_group)
        row[GROUP_COLUMN] = latest_group
        row["전체 변화"] = row[snapshots[-1]["label"]] - row[snapshots[0]["label"]]
        rows.append(row)

    rows.sort(key=lambda value: value["전체 변화"], reverse=True)
    return rows


def add_table(ws, headers, rows, start_row=1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3F5F73")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, row.get(header, ""))

    last_row = start_row + len(rows)
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = f"A{start_row + 1}"
    autosize_columns(ws)
    return last_row


def autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 40))
        ws.column_dimensions[letter].width = max(10, max_length + 2)


def apply_number_formats(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = '#,##0'


def apply_delta_colors(ws, headers, first_data_row=2):
    green = PatternFill("solid", fgColor="D9EAD3")
    red = PatternFill("solid", fgColor="F4CCCC")
    blue = PatternFill("solid", fgColor="CFE2F3")
    orange = PatternFill("solid", fgColor="FCE5CD")

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        data_range = f"{letter}{first_data_row}:{letter}{ws.max_row}"
        if header.endswith(" 변화"):
            ws.conditional_formatting.add(data_range, CellIsRule(operator="greaterThan", formula=["0"], fill=green))
            ws.conditional_formatting.add(data_range, CellIsRule(operator="lessThan", formula=["0"], fill=red))

    if "상태" in headers:
        status_col = get_column_letter(headers.index("상태") + 1)
        for row_idx in range(first_data_row, ws.max_row + 1):
            status = ws[f"{status_col}{row_idx}"].value
            if status == "신규":
                fill = blue
            elif status == "이탈":
                fill = orange
            else:
                continue
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row_idx, col_idx).fill = fill


def add_summary_sheet(wb, snapshots, comparison, group_summary, top_n):
    ws = wb.active
    ws.title = "요약"

    current = snapshots[-1]
    previous = snapshots[-2]
    maintained = sum(1 for row in comparison if row["상태"] == "유지")
    added = sum(1 for row in comparison if row["상태"] == "신규")
    removed = sum(1 for row in comparison if row["상태"] == "이탈")

    rows = [
        ("이전 파일", previous["path"].name),
        ("현재 파일", current["path"].name),
        ("이전 시점", previous["label"]),
        ("현재 시점", current["label"]),
        ("유지 인원", maintained),
        ("신규 인원", added),
        ("이탈 인원", removed),
    ]

    for row_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row_idx, 1, label).font = Font(bold=True)
        ws.cell(row_idx, 2, value)

    start = len(rows) + 3
    ws.cell(start, 1, "주요 변화").font = Font(bold=True, size=14)
    metric_rows = []
    for column in ["세력치", "공헌총량", "전공총량", "협공총량", "기부총량"]:
        metric_rows.append(
            {
                "항목": column,
                "전체 변화": sum(row[f"{column} 변화"] for row in comparison),
                "상승 인원": sum(1 for row in comparison if row[f"{column} 변화"] > 0),
                "하락 인원": sum(1 for row in comparison if row[f"{column} 변화"] < 0),
            }
        )
    add_table(ws, ["항목", "전체 변화", "상승 인원", "하락 인원"], metric_rows, start + 1)

    group_start = start + len(metric_rows) + 5
    ws.cell(group_start, 1, "그룹별 세력치 변화 Top").font = Font(bold=True, size=14)
    chart_rows = group_summary[: min(top_n, len(group_summary))]
    headers = ["그룹", "세력치 변화", "공헌총량 변화", "전공총량 변화", "신규", "이탈"]
    add_table(ws, headers, chart_rows, group_start + 1)

    if chart_rows:
        chart = BarChart()
        chart.title = "그룹별 세력치 변화"
        chart.y_axis.title = "변화량"
        chart.x_axis.title = "그룹"
        data = Reference(ws, min_col=2, min_row=group_start + 1, max_row=group_start + len(chart_rows) + 1)
        categories = Reference(ws, min_col=1, min_row=group_start + 2, max_row=group_start + len(chart_rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "H12")

    apply_number_formats(ws)
    autosize_columns(ws)


def add_top_sheet(wb, comparison, top_n):
    ws = wb.create_sheet("변화량 Top")
    current_row = 1
    headers = ["회원", "그룹", "상태", "이전값", "현재값", "변화"]

    for column in TOP_COLUMNS:
        for direction, reverse in [("상승", True), ("하락", False)]:
            title = f"{column} {direction} Top {top_n}"
            ws.cell(current_row, 1, title).font = Font(bold=True, size=14)
            current_row += 1
            metric = f"{column} 변화"
            rows = sorted(comparison, key=lambda item: item[metric], reverse=reverse)
            if direction == "상승":
                rows = [row for row in rows if row[metric] > 0]
            else:
                rows = [row for row in rows if row[metric] < 0]
            table_rows = [
                {
                    "회원": row[KEY_COLUMN],
                    "그룹": row[GROUP_COLUMN],
                    "상태": row["상태"],
                    "이전값": row[f"이전 {column}"],
                    "현재값": row[f"현재 {column}"],
                    "변화": row[metric],
                }
                for row in rows[:top_n]
            ]
            add_table(ws, headers, table_rows, current_row)
            current_row += max(len(table_rows), 1) + 3

    apply_number_formats(ws)
    apply_delta_colors(ws, ["회원", "그룹", "상태", "이전값", "현재값", "변화"], 2)
    autosize_columns(ws)


def add_metric_sheet(wb, comparison, sheet_title, source_column, use_active=False):
    ws = wb.active if use_active else wb.create_sheet(sheet_title)
    ws.title = sheet_title

    headers = ["상태", KEY_COLUMN, GROUP_COLUMN, "이전값", "현재값", "변화"]
    rows = []
    for row in comparison:
        rows.append(
            {
                "상태": row["상태"],
                KEY_COLUMN: row[KEY_COLUMN],
                GROUP_COLUMN: row[GROUP_COLUMN],
                "이전값": row[f"이전 {source_column}"],
                "현재값": row[f"현재 {source_column}"],
                "변화": row[f"{source_column} 변화"],
            }
        )

    rows.sort(key=lambda item: item["변화"], reverse=True)
    add_table(ws, headers, rows)
    apply_number_formats(ws)
    apply_delta_colors(ws, headers)


def build_metric_trend_rows(snapshots, source_column, member_filter=None):
    if member_filter:
        members = sorted(member_filter)
    else:
        members = sorted({member for snapshot in snapshots for member in snapshot["rows"]})
    rows = []

    for member in members:
        latest_group = ""
        values = []
        for snapshot in snapshots:
            source = snapshot["rows"].get(member)
            value = source.get(source_column, 0) if source else 0
            values.append(value)
            if source:
                latest_group = source.get(GROUP_COLUMN, latest_group)

        row = {
            KEY_COLUMN: member,
            GROUP_COLUMN: latest_group,
            "처음값": values[0],
            "마지막값": values[-1],
            "전체 변화": values[-1] - values[0],
        }
        for snapshot, value in zip(snapshots, values):
            row[snapshot["label"]] = value
        rows.append(row)

    rows.sort(key=lambda item: item["전체 변화"], reverse=True)
    return rows


def set_chart_cell_anchor(chart, start_col, start_row, end_col, end_row):
    chart.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=start_col - 1, row=start_row - 1),
        to=AnchorMarker(col=end_col - 1, row=end_row - 1),
    )


def add_line_chart_for_rows(ws, title, row_indexes, date_start_col, date_end_col, anchor_row):
    if not row_indexes:
        return

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "수치"
    chart.x_axis.title = "날짜"
    chart.height = 22
    chart.width = 42

    categories = Reference(ws, min_col=date_start_col, max_col=date_end_col, min_row=1)
    for row_idx in row_indexes:
        values = Reference(ws, min_col=date_start_col, max_col=date_end_col, min_row=row_idx)
        series = Series(values, title=str(ws.cell(row_idx, 1).value))
        chart.series.append(series)

    chart.set_categories(categories)
    set_chart_cell_anchor(chart, 1, anchor_row, 24, anchor_row + 34)
    ws.add_chart(chart)


def add_trend_sheet(wb, snapshots, sheet_title, source_column, top_n, member_filter=None):
    ws = wb.create_sheet(sheet_title)
    rows = build_metric_trend_rows(snapshots, source_column, member_filter)
    date_headers = [snapshot["label"] for snapshot in snapshots]
    headers = [KEY_COLUMN, GROUP_COLUMN, "처음값", "마지막값", "전체 변화"] + date_headers

    add_table(ws, headers, rows)
    apply_number_formats(ws)
    apply_delta_colors(ws, headers)

    date_start_col = len(headers) - len(date_headers) + 1
    date_end_col = len(headers)
    chart_rows = list(range(2, min(top_n, len(rows)) + 2))
    chart_title = f"{source_column} 전체 기간 {len(chart_rows)}명"
    chart_anchor_row = max(len(rows) + 4, 8)
    add_line_chart_for_rows(ws, chart_title, chart_rows, date_start_col, date_end_col, chart_anchor_row)


def add_member_chart_sheet(wb, snapshots, member):
    ws = wb.create_sheet("개인 그래프")
    metrics = [
        ("세력치", "세력치"),
        ("공헌", "공헌총량"),
        ("전공", "전공총량"),
        ("협공", "협공총량"),
        ("기부", "기부총량"),
    ]

    headers = ["날짜"] + [label for label, _ in metrics]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3F5F73")
        cell.alignment = Alignment(horizontal="center")

    found = False
    for row_idx, snapshot in enumerate(snapshots, start=2):
        source = snapshot["rows"].get(member)
        if source:
            found = True
        ws.cell(row_idx, 1, snapshot["label"])
        for col_idx, (_, source_column) in enumerate(metrics, start=2):
            ws.cell(row_idx, col_idx, source.get(source_column, 0) if source else 0)

    apply_number_formats(ws)
    autosize_columns(ws)

    chart = LineChart()
    chart.title = f"{member} 개인 변화"
    chart.y_axis.title = "수치"
    chart.x_axis.title = "날짜"
    chart.height = 20
    chart.width = 38
    data = Reference(ws, min_col=2, max_col=len(headers), min_row=1, max_row=len(snapshots) + 1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(snapshots) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    set_chart_cell_anchor(chart, 1, len(snapshots) + 4, 20, len(snapshots) + 34)
    ws.add_chart(chart)

    if not found:
        ws.cell(len(snapshots) + 3, 1, f"'{member}' 회원을 CSV에서 찾지 못했습니다.")


def add_workbook_sheets(wb, snapshots, comparison, top_n, member=None, member_filter=None):
    chart_count = len(comparison) if member_filter else top_n

    metric_sheets = [
        ("세력치 변화", "세력치"),
        ("공헌 변화", "공헌총량"),
        ("전공 변화", "전공총량"),
        ("협공 변화", "협공총량"),
        ("기부 변화", "기부총량"),
        ("랭킹 변화", "공헌 랭킹"),
    ]

    for index, (sheet_title, source_column) in enumerate(metric_sheets):
        add_metric_sheet(wb, comparison, sheet_title, source_column, use_active=(index == 0))

    trend_sheets = [
        ("세력치 추이", "세력치"),
        ("공헌 추이", "공헌총량"),
        ("전공 추이", "전공총량"),
    ]
    for sheet_title, source_column in trend_sheets:
        add_trend_sheet(wb, snapshots, sheet_title, source_column, chart_count, member_filter)

    if member:
        add_member_chart_sheet(wb, snapshots, member)


def load_member_filter(path):
    if not path.exists():
        return None

    members = set()
    with path.open("r", encoding="utf-8-sig") as handle:
        for line in handle:
            value = line.strip()
            if not value or value.startswith("#"):
                continue
            members.add(value)

    return members or None


def main():
    parser = argparse.ArgumentParser(description="동맹 통계 CSV 변화량 리포트를 생성합니다.")
    parser.add_argument("--input-dir", default=DEFAULT_INPUT_DIR, help="CSV 파일들이 있는 폴더")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="생성할 xlsx 파일 경로")
    parser.add_argument("--top", type=int, default=10, help="Top N 개수")
    parser.add_argument("--member", default="", help="개인 그래프를 만들 회원명")
    parser.add_argument("--members-file", default=DEFAULT_MEMBERS_FILE, help="리포트에 포함할 회원명 목록 txt")
    parser.add_argument("--all-members", action="store_true", help="members.txt를 무시하고 CSV 전체 회원을 출력")
    parser.add_argument("--no-prompt", action="store_true", help="실행 중 회원명 입력을 묻지 않음")
    args = parser.parse_args()

    if getattr(sys, "frozen", False):
        script_dir = Path(sys.executable).resolve().parent
    else:
        script_dir = Path(__file__).resolve().parent
    members_file = Path(args.members_file)
    if not members_file.is_absolute():
        members_file = script_dir / members_file
    member_filter = None if args.all_members else load_member_filter(members_file)

    interactive = not args.no_prompt
    member = args.member.strip()
    if interactive and not member and not member_filter:
        try:
            member = input("개인 그래프를 만들 회원명 입력(건너뛰려면 Enter): ").strip()
        except EOFError:
            member = ""

    input_dir = Path(args.input_dir)
    if not input_dir.is_absolute():
        input_dir = script_dir / input_dir

    snapshots = load_snapshots(input_dir)
    comparison = build_comparison(snapshots[-2], snapshots[-1], member_filter)

    wb = Workbook()
    add_workbook_sheets(wb, snapshots, comparison, args.top, member or None, member_filter)

    output = Path(args.output)
    if not output.is_absolute():
        if output.parent == Path("."):
            output = script_dir / DEFAULT_OUTPUT_DIR / output
        else:
            output = script_dir / output
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    print(f"생성 완료: {output.resolve()}")
    print(f"비교 기준: {snapshots[-2]['path'].name} -> {snapshots[-1]['path'].name}")
    print(f"CSV 파일 수: {len(snapshots)}")
    if member_filter:
        print(f"회원 필터: {members_file.resolve()} ({len(member_filter)}명)")
    if member:
        print(f"개인 그래프: {member}")

    if interactive:
        input("완료되었습니다. 창을 닫으려면 Enter를 누르세요.")


if __name__ == "__main__":
    main()
